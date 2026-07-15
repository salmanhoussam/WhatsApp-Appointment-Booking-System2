"""
Product-catalog import — concrete and product-specific, per approved
Phase 2.5 decisions (.claudelocaldocs/phase2.5-data-validation-report.md).

This is deliberately NOT a generic "import any file" engine. It knows the
exact 13-column shape verified against a real POS price-list export
(category, name, code, tax, unit, currency, price, price+tax, activation
date, min price, min price+tax, created by, created at), skips the 3
header rows, and normalizes whitespace only — it does not split `unit`,
does not derive/drop the *_incl_tax columns, and does not build price
history. See .claude/rules/team-roles.md for the rule on when a shared
Import Engine would become justified (not yet — this is import type #1).
"""

import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from plugins import plugin_manager

_HEADER_ROWS_TO_SKIP = 3


def _normalize_text(value) -> Optional[str]:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value).strip())
    return text or None


def _cell_to_str(value) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _row_to_item(cell) -> dict:
    """`cell(index)` returns the already-Python-typed value for that column."""
    return {
        "category": _normalize_text(cell(0)),
        "name": _normalize_text(cell(1)),
        "sku_code": _cell_to_str(cell(2)),
        "tax_class": _normalize_text(cell(3)),
        "unit": _normalize_text(cell(4)),  # kept whole — decision 6, no splitting yet
        "currency": _normalize_text(cell(5)),
        "price": cell(6),
        "price_incl_tax": cell(7),  # stored as-given, not derived — decision/architectural directive
        "activation_date": cell(8),  # metadata only — decision 7
        "min_price": cell(9),  # stored if present, unenforced — decision 5
        "min_price_incl_tax": cell(10),
        "source_created_by": _normalize_text(cell(11)),
        "source_created_at": cell(12),
    }


def _parse_xls(path: Path) -> list[dict]:
    import xlrd

    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(0)
    items = []
    for r in range(_HEADER_ROWS_TO_SKIP, sheet.nrows):
        if sheet.cell(r, 1).value in (None, ""):
            continue

        def cell(c, _r=r):
            cl = sheet.cell(_r, c)
            if cl.ctype == xlrd.XL_CELL_DATE:
                return xlrd.xldate.xldate_as_datetime(cl.value, book.datemode).isoformat()
            if cl.ctype == xlrd.XL_CELL_EMPTY or cl.value == "":
                return None
            return cl.value

        items.append(_row_to_item(cell))
    return items


def _parse_xlsx(path: Path) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), data_only=True)
    sheet = wb.worksheets[0]
    items = []
    for row in sheet.iter_rows(min_row=_HEADER_ROWS_TO_SKIP + 1, values_only=True):
        if not row or row[1] in (None, ""):
            continue

        def cell(c, _row=row):
            v = _row[c] if c < len(_row) else None
            return v.isoformat() if isinstance(v, datetime) else v

        items.append(_row_to_item(cell))
    return items


def _parse(path: Path) -> list[dict]:
    if not path.exists():
        raise ValueError(f"File not found: {path}")
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return _parse_xls(path)
    if suffix == ".xlsx":
        return _parse_xlsx(path)
    raise ValueError(f"Unsupported catalog file type '{suffix}' — only .xls and .xlsx are supported")


def _find_possible_duplicates(items: list[dict]) -> list[dict]:
    """Report-only. Never used to filter or remove rows — decision 3."""
    seen: dict[tuple, str] = {}
    duplicates = []
    for item in items:
        key = (item.get("name"), item.get("price"), item.get("unit"))
        if key in seen:
            duplicates.append({
                "name": item.get("name"), "price": item.get("price"), "unit": item.get("unit"),
                "sku_codes": [seen[key], item.get("sku_code")],
            })
        else:
            seen[key] = item.get("sku_code")
    return duplicates


def preview_catalog_import(file_path: str) -> dict:
    """Dry run — parses and reports what would happen. Writes nothing."""
    items = _parse(Path(file_path))
    return {
        "source_file": Path(file_path).name,
        "row_count": len(items),
        "possible_duplicates": _find_possible_duplicates(items),
        "sample": items[:3],
    }


def import_product_catalog(file_path: str) -> dict:
    """Parses and commits: one import_batches row + all product_catalog rows, atomically."""
    path = Path(file_path)
    items = _parse(path)
    if not items:
        raise ValueError(f"No product rows found in {file_path}")

    duplicates = _find_possible_duplicates(items)
    batch_id = str(uuid.uuid4())
    result = plugin_manager.execute("commit_product_catalog_import", {
        "batch_id": batch_id,
        "source_file": path.name,
        "imported_at": datetime.now(timezone.utc).isoformat(),
        "items": items,
    })
    if "error" in result:
        raise ValueError(result["error"])

    return {
        "batch_id": batch_id,
        "source_file": path.name,
        "items_inserted": result["items_inserted"],
        "possible_duplicates": duplicates,
    }


def list_catalog_imports() -> list[dict]:
    result = plugin_manager.execute("list_import_batches", {"entity_type": "product_catalog"})
    return result.get("batches", [])


def search_catalog(query: str) -> list[dict]:
    if not query or not query.strip():
        raise ValueError("Search query is required")
    result = plugin_manager.execute("search_product_catalog", {"query": query.strip()})
    return result.get("items", [])


def rollback_catalog_import(batch_id: str) -> dict:
    result = plugin_manager.execute("rollback_import_batch", {"batch_id": batch_id})
    if "error" in result:
        raise ValueError(result["error"])
    return result
